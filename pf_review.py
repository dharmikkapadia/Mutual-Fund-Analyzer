"""
pf_review.py — the monthly Mutual-Fund Portfolio Review engine.

Replicates the hand-built "MF Portfolio Review" workbook: one row per
scheme carrying the invested value plus fund parameters (P/B, P/E, AUM,
large/mid/small-cap split, debt & cash, 19 sector weights), and a summary
row where every parameter is the investment-value-weighted average
(Excel's SUMPRODUCT(param, values)/total).

Pure pandas/openpyxl — no network, no UI — so all the maths and the Excel
export are unit-testable offline. Percentages are handled on a 0–100 scale
throughout this module and only converted to fractions inside the Excel
export (the workbook stores 0.7675 and formats it as 76.75%).
"""

from __future__ import annotations

import io
import json
import re

import numpy as np
import pandas as pd

from vr_data import SECTOR_COLS

# Display/spreadsheet column titles, in the workbook's exact order
CAP_COLS = ["Large Stocks", "Mid cap Stocks", "Small cap Stocks"]
VALUE_COL = "Value"
PARAM_COLS = ["P/B", "P/E", "Aum in cr", *CAP_COLS, "Debt & Cash",
              *SECTOR_COLS]
# parameters that are % of the portfolio (drive the Total check column)
PCT_COLS = [*CAP_COLS, "Debt & Cash", *SECTOR_COLS]
TOTAL_COL = "Total"          # Debt & Cash + all sectors ≈ 100

# vr_data record key -> column title
_KEY_TO_COL = {"pb": "P/B", "pe": "P/E", "aum_cr": "Aum in cr",
               "large": "Large Stocks", "mid": "Mid cap Stocks",
               "small": "Small cap Stocks", "debt_cash": "Debt & Cash"}


def params_to_row(params: dict | None) -> dict:
    """Flatten a vr_data-style record into {column title: value}."""
    row = {c: np.nan for c in PARAM_COLS}
    if params:
        for key, col in _KEY_TO_COL.items():
            v = params.get(key)
            row[col] = float(v) if v is not None else np.nan
        for sector, v in (params.get("sectors") or {}).items():
            if sector in row and v is not None:
                row[sector] = float(v)
        # a parsed sector table means absent sectors are genuinely 0%
        if params.get("sectors"):
            for c in SECTOR_COLS:
                if pd.isna(row[c]):
                    row[c] = 0.0
    return row


def review_frame(rows: list[dict]) -> pd.DataFrame:
    """The review table: one row per scheme, workbook column order.

    `rows` items: {"name": str, "value": float, **{column: value}} — i.e.
    already flattened via params_to_row (or hand-edited in the UI grid).
    """
    recs = []
    for r in rows:
        v = pd.to_numeric(r.get("value"), errors="coerce")
        rec = {"Scheme Name": r.get("name", ""),
               VALUE_COL: 0.0 if pd.isna(v) else float(v)}
        for c in PARAM_COLS:
            rec[c] = r.get(c, np.nan)
        recs.append(rec)
    df = pd.DataFrame(recs, columns=["Scheme Name", VALUE_COL, *PARAM_COLS])
    for c in [VALUE_COL, *PARAM_COLS]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df[TOTAL_COL] = df[["Debt & Cash", *SECTOR_COLS]].sum(
        axis=1, min_count=1)
    return df


def weighted_summary(df: pd.DataFrame) -> pd.Series:
    """Value-weighted average of every parameter (the workbook's row 9).

    Where the workbook's SUMPRODUCT treats a blank as 0 against the full
    total, here a scheme missing a parameter is excluded from both the
    numerator and denominator of *that* parameter, so partial data doesn't
    silently drag averages toward zero.
    """
    out = {VALUE_COL: float(df[VALUE_COL].sum())}
    for c in PARAM_COLS:
        vals, w = df[c], df[VALUE_COL]
        mask = vals.notna() & w.notna() & (w > 0)
        denom = float(w[mask].sum())
        out[c] = (float((vals[mask] * w[mask]).sum()) / denom
                  if denom > 0 else np.nan)
    s = pd.Series(out)
    s[TOTAL_COL] = s[["Debt & Cash", *SECTOR_COLS]].sum()
    return s


# --------------------------------------------------------------------------- #
# Monthly snapshots
# --------------------------------------------------------------------------- #
def snapshot_pack(rows: list[dict], as_on: str) -> dict:
    """JSON-able snapshot of the review grid for one month."""
    clean = []
    for r in rows:
        rec = {"code": r.get("code"), "name": r.get("name", ""),
               "value": _num(r.get("value")), "vr_url": r.get("vr_url")}
        for c in PARAM_COLS:
            rec[c] = _num(r.get(c))
        clean.append(rec)
    return {"as_on": str(as_on), "rows": clean}


def snapshot_rows(snap: dict) -> list[dict]:
    """Back from snapshot_pack to review_frame-ready rows."""
    return [dict(r) for r in (snap or {}).get("rows", [])]


def _num(v):
    try:
        f = float(v)
        return None if np.isnan(f) else f
    except (TypeError, ValueError):
        return None


def snapshots_to_json(snaps: dict) -> str:
    return json.dumps(snaps, separators=(",", ":"))


def snapshots_from_json(raw: str) -> dict:
    try:
        obj = json.loads(raw or "{}")
        return obj if isinstance(obj, dict) else {}
    except ValueError:
        return {}


# --------------------------------------------------------------------------- #
# Workbook import — a manually maintained review .xlsx becomes a snapshot
# --------------------------------------------------------------------------- #
def _norm_title(s) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


# normalised header -> canonical field; sectors and the rest are appended
_TITLE_MAP = {"scheme name": "__name", "p/b": "P/B", "p/e": "P/E",
              "debt & cash": "Debt & Cash", "large stocks": "Large Stocks",
              "mid cap stocks": "Mid cap Stocks",
              "small cap stocks": "Small cap Stocks"}
_TITLE_MAP.update({_norm_title(c): c for c in SECTOR_COLS})


def _title_field(title: str) -> str | None:
    t = _norm_title(title)
    if t in _TITLE_MAP:
        return _TITLE_MAP[t]
    if t.startswith("aum"):
        return "Aum in cr"
    if t.startswith("value"):
        return "__value"
    return None


def parse_workbook(data: bytes) -> dict:
    """Parse a manual 'MF Portfolio Review' workbook into snapshot rows.

    Header-driven and tolerant: finds the sheet and row containing
    'Scheme Name', maps columns by (normalised) title, reads scheme rows
    until the summary row (blank scheme name), and rescales
    fraction-stored percentages (0.7675 shown as 76.75%) to the 0-100
    scale the app uses. Returns {"as_on": str|None, "rows": [...]}.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True)
    header_row = cols = ws_found = None
    for ws in wb.worksheets:
        for r in range(1, 6):
            titles = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if any(_norm_title(t) == "scheme name" for t in titles):
                header_row, ws_found = r, ws
                cols = {c: _title_field(t)
                        for c, t in enumerate(titles, start=1)
                        if _title_field(t) is not None}
                break
        if cols:
            break
    if not cols:
        raise ValueError("no 'Scheme Name' header found — is this the "
                         "review-workbook layout?")

    as_on = None
    for c, f in cols.items():
        if f == "__value":
            m = re.search(r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
                          str(ws_found.cell(header_row, c).value or ""))
            if m:
                as_on = m.group(1)

    rows = []
    for r in range(header_row + 1, ws_found.max_row + 1):
        rec = {c2: None for c2 in PARAM_COLS}
        name = value = None
        for c, f in cols.items():
            v = ws_found.cell(r, c).value
            if f == "__name":
                name = str(v).strip() if v not in (None, "") else None
            elif f == "__value":
                value = _num(v)
            else:
                rec[f] = _num(v)
        if not name:            # summary row / end of table
            break
        rows.append({"name": name, "value": value, "code": None,
                     "vr_url": None, **rec})

    # fraction-stored percentages -> 0-100 (their max would be ~1.0)
    pct_vals = [r[c] for r in rows for c in PCT_COLS if r[c] is not None]
    if pct_vals and max(pct_vals) <= 1.5:
        for r in rows:
            for c in PCT_COLS:
                if r[c] is not None:
                    r[c] = round(r[c] * 100.0, 6)
    return {"as_on": as_on, "rows": rows}


def month_key_from_as_on(as_on) -> str | None:
    """'31/03/2026' -> '2026-03' (regex-based, tolerates odd dates)."""
    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", str(as_on or ""))
    if not m:
        return None
    month, year = int(m.group(2)), int(m.group(3))
    if year < 100:
        year += 2000
    if not 1 <= month <= 12:
        return None
    return f"{year:04d}-{month:02d}"


# --------------------------------------------------------------------------- #
# Excel export — the uploaded workbook's exact layout, with live formulas
# --------------------------------------------------------------------------- #
def to_excel_bytes(df: pd.DataFrame, as_on: str,
                   sheet_name: str = "PF Review") -> bytes:
    """Render the review as .xlsx bytes in the source workbook's layout.

    Row 1 headers, one row per scheme, then the summary row: total value
    via SUM, every parameter via SUMPRODUCT(param, values)/total — live
    formulas, so the sheet keeps recalculating if edited by hand.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = ["Sr.No.", "Scheme Name", f"Value as on {as_on}",
               "Weight in PF", "P/B", "P/E", "Aum in cr", *CAP_COLS,
               "Debt & Cash", *SECTOR_COLS, TOTAL_COL]
    bold = Font(bold=True)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = bold

    n = len(df)
    first, last, total_row = 2, n + 1, n + 2
    # column letters: A Sr, B name, C value, D weight, E P/B, F P/E, G AUM,
    # H..J caps, K debt&cash, L.. sectors, last = Total
    kcol, last_col = 11, 11 + len(SECTOR_COLS) + 1
    val_fmt = '_ * #,##0_ ;_ * \\-#,##0_ ;_ * "-"??_ ;_ @_ '
    pct2 = "0.00%"

    for i, (_, r) in enumerate(df.iterrows()):
        row = first + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=str(r["Scheme Name"]))
        ws.cell(row=row, column=3,
                value=_x(r[VALUE_COL])).number_format = val_fmt
        d = ws.cell(row=row, column=4, value=f"=C{row}/C${total_row}")
        d.number_format = "0%"
        ws.cell(row=row, column=5, value=_x(r["P/B"]))
        ws.cell(row=row, column=6, value=_x(r["P/E"]))
        ws.cell(row=row, column=7,
                value=_x(r["Aum in cr"])).number_format = val_fmt
        for k, col in enumerate([*CAP_COLS, "Debt & Cash", *SECTOR_COLS]):
            c = ws.cell(row=row, column=8 + k, value=_frac(r[col]))
            c.number_format = pct2
        t = ws.cell(row=row, column=last_col,
                    value=f"=SUM({get_column_letter(kcol)}{row}:"
                          f"{get_column_letter(last_col - 1)}{row})")
        t.number_format = pct2
        t.font = bold

    # summary row — the workbook's SUMPRODUCT weighted averages
    ws.cell(row=total_row, column=3,
            value=f"=SUM(C{first}:C{last})").number_format = "#,##0"
    ws.cell(row=total_row, column=3).font = bold
    for j in range(5, last_col):
        letter = get_column_letter(j)
        c = ws.cell(
            row=total_row, column=j,
            value=(f"=SUMPRODUCT({letter}{first}:{letter}{last},"
                   f"$C${first}:$C${last})/$C${total_row}"))
        c.font = bold
        c.number_format = "0.00" if j in (5, 6) else (
            "#,##0" if j == 7 else pct2)

    widths = {1: 6.4, 2: 33.0, 3: 21.7, 4: 12.3, 5: 6.5, 6: 6.5, 7: 10.0}
    for j in range(1, last_col + 1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(j, 11.7)
    ws.freeze_panes = "C2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _x(v):
    """NaN -> empty cell."""
    return None if v is None or (isinstance(v, float) and np.isnan(v)) else v


def _frac(v):
    """Percent (0-100) -> stored fraction, like the source workbook."""
    v = _x(v)
    return None if v is None else round(float(v) / 100.0, 6)
